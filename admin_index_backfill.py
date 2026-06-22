"""
admin_index_backfill.py — Scorr
================================
One-time (repeatable / idempotent) backfill of the three Indian indices that
were missing from raw_prices: SENSEX, FINNIFTY, MIDCAPNIFTY. Daily OHLC pulled
from the Yahoo chart API, default 5-year window, UPSERT into raw_prices with the
SAME schema + same fetch/parse path as NIFTY50 / BANKNIFTY (so the rows line up
exactly with the existing index history).

Route
-----
POST /api/admin/backfill-indian-indices?lookback=5y&dry_run=false
  ADMIN_TOKEN required (X-Admin-Token header).
  dry_run=true  -> test-fetch only, report per-symbol row counts, NO DB write.
  dry_run=false -> fetch + UPSERT into raw_prices.

GET /api/admin/indices-excel
  Public download of a 5-sheet 5yr daily-OHLC .xlsx straight from raw_prices.

Per symbol it test-fetches first and logs the row count. FINNIFTY tries a
candidate ticker list (^CNXFINANCE first per the task spec, falling back to
NIFTY_FIN_SERVICE.NS) and reports which ticker actually returned data.

Nightly freshness
-----------------
After a successful backfill these symbols are auto-included in the nightly
yahoo_daily_update run (it pulls `SELECT DISTINCT symbol FROM raw_prices`); the
matching Yahoo ticker overrides live in yahoo_daily_update.INDICES.

Note on "Midcap Nifty": Yahoo has no clean feed for the Nifty Midcap Select
(the F&O index); ^NSEMDCP50 (Nifty Midcap 50) is used as the closest proxy, per
the task spec.
"""

import os
import asyncio
import logging

from fastapi import APIRouter, HTTPException, Header

import yahoo_daily_update as ydu

log = logging.getLogger("scorr.idx_backfill")
router = APIRouter(prefix="/api/admin", tags=["admin"])

# store_as symbol -> ordered Yahoo tickers to try (first that returns data wins)
INDEX_TICKERS = {
    "SENSEX":      ["^BSESN"],
    "FINNIFTY":    ["^CNXFINANCE", "NIFTY_FIN_SERVICE.NS"],
    "MIDCAPNIFTY": ["^NSEMDCP50"],
}

# Excel export: (sheet title, raw_prices symbol) in display order
EXCEL_SHEETS = [
    ("Nifty 50",     "NIFTY50"),
    ("Bank Nifty",   "BANKNIFTY"),
    ("FinNifty",     "FINNIFTY"),
    ("Sensex",       "SENSEX"),
    ("Midcap Nifty", "MIDCAPNIFTY"),
]


@router.get("/indices-excel")
def indices_excel():
    """Download a 5-sheet .xlsx of 5yr daily OHLC for all 5 indices straight
    from raw_prices. Public (index OHLC is public market data) so it opens with
    a plain browser click. openpyxl is imported lazily so a missing dep can
    never break app boot."""
    import io
    from datetime import datetime
    import psycopg
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    counts = {}
    with psycopg.connect(ydu.DB_URL) as conn, conn.cursor() as cur:
        for title, sym in EXCEL_SHEETS:
            ws = wb.create_sheet(title=title)
            ws.append(["Date", "Open", "High", "Low", "Close", "Volume"])
            for c in ws[1]:
                c.font = Font(bold=True)
            cur.execute(
                "SELECT price_date, open, high, low, close, volume "
                "FROM raw_prices WHERE symbol=%s ORDER BY price_date", (sym,))
            n = 0
            for d, o, h, l, cl, v in cur.fetchall():
                ws.append([
                    d,
                    float(o) if o is not None else None,
                    float(h) if h is not None else None,
                    float(l) if l is not None else None,
                    float(cl) if cl is not None else None,
                    int(v) if v is not None else None,
                ])
                n += 1
            counts[sym] = n
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 12
            for col in ("B", "C", "D", "E", "F"):
                ws.column_dimensions[col].width = 13
            for (cell,) in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                cell.number_format = "yyyy-mm-dd"
    log.info(f"indices_excel: {counts}")
    buf = io.BytesIO()
    wb.save(buf)
    fn = f"Indices_5yr_OHLC_{datetime.now().strftime('%Y%m%d')}.xlsx"
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


def run_backfill(lookback: str = "5y", dry_run: bool = False) -> dict:
    """Backfill SENSEX / FINNIFTY / MIDCAPNIFTY daily OHLC into raw_prices.

    Pure callable (no auth) so it can run either via the FastAPI route below or
    in-process from the MCP dispatch layer. Per symbol: test-fetch, log row
    count, then UPSERT (unless dry_run). FINNIFTY walks its candidate ticker
    list until one returns data and reports which one worked."""
    ydu.LOOKBACK = lookback  # _fetch_symbol builds the chart URL from this
    results = {}

    for store_as, candidates in INDEX_TICKERS.items():
        chosen, rows = None, []
        for tk in candidates:
            try:
                rows = _fetch_rows(store_as, tk)
            except Exception as e:
                log.warning(f"backfill {store_as} via {tk}: {e}")
                rows = []
            if rows:
                chosen = tk
                break

        if not rows:
            results[store_as] = {"status": "no_data", "tried": candidates}
            log.warning(f"backfill {store_as}: no data from {candidates}")
            continue

        first_d, last_d = rows[0][1], rows[-1][1]
        upserted = 0
        if not dry_run:
            upserted = ydu._commit_rows({store_as: rows})

        results[store_as] = {
            "status": "ok",
            "ticker": chosen,
            "rows": len(rows),
            "first_date": str(first_d),
            "last_date": str(last_d),
            "upserted": upserted,
            "dry_run": dry_run,
        }
        log.info(f"backfill {store_as} [{chosen}]: {len(rows)} rows "
                 f"{first_d}..{last_d} dry_run={dry_run}")

    return {"status": "done", "lookback": lookback, "dry_run": dry_run,
            "results": results}


def _fetch_rows(store_as: str, ticker: str):
    """Fetch daily OHLC rows for ONE ticker (no DB write).

    Reuses yahoo_daily_update's exact fetch/parse path by temporarily pointing
    its INDICES override at `ticker`, so the resulting rows are byte-for-byte
    consistent with how NIFTY50/BANKNIFTY are stored. Returns the list of
    UPSERT-ready tuples (possibly empty)."""
    prev = ydu.INDICES.get(store_as)
    ydu.INDICES[store_as] = ticker
    try:
        got, _failed = asyncio.run(
            ydu._run_pass([store_as], ydu.RETRY_SEMAPHORE, ydu.RETRY_SLEEP)
        )
    finally:
        if prev is None:
            ydu.INDICES.pop(store_as, None)
        else:
            ydu.INDICES[store_as] = prev
    return got.get(store_as, [])


@router.post("/backfill-indian-indices")
def backfill_indian_indices(lookback: str = "5y", dry_run: bool = False,
                            x_admin_token: str = Header(None)):
    """HTTP wrapper around run_backfill (ADMIN_TOKEN). Available once the router
    is wired in main.py; the MCP path calls run_backfill() in-process."""
    if x_admin_token != os.getenv("ADMIN_TOKEN"):
        raise HTTPException(401, "Unauthorized")
    return run_backfill(lookback=lookback, dry_run=dry_run)
